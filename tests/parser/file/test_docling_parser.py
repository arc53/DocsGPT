"""Comprehensive tests for application/parser/file/docling_parser.py

Covers: DoclingParser (init, _init_parser, OCR engine selection, _export_content,
parse_file), subclass initialization, error handling.
"""

import logging
import os
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest


# =====================================================================
# DoclingParser - Init
# =====================================================================


@pytest.mark.unit
class TestDoclingParserInit:

    def test_default_init(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()
        assert parser.ocr_enabled is True
        assert parser.table_structure is True
        assert parser.export_format == "markdown"
        assert parser.ocr_engine is None  # None -> settings.OCR_ENGINE at build
        assert parser.ocr_languages is None  # None -> the engine's own default
        assert parser.force_full_page_ocr is False
        assert parser._converter is None

    def test_custom_init(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(
            ocr_enabled=False,
            table_structure=False,
            export_format="text",
            ocr_engine="rapidocr",
            ocr_languages=["german"],
            force_full_page_ocr=True,
        )
        assert parser.ocr_enabled is False
        assert parser.table_structure is False
        assert parser.export_format == "text"
        assert parser.ocr_engine == "rapidocr"
        assert parser.ocr_languages == ["german"]
        assert parser.force_full_page_ocr is True


# =====================================================================
# Init Parser
# =====================================================================


@pytest.mark.unit
class TestDoclingParserInitParser:

    def test_init_parser_raises_without_docling(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()

        with patch("importlib.util.find_spec", return_value=None):
            with pytest.raises(ImportError, match="docling is required"):
                parser._init_parser()

    def test_init_parser_success(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()

        mock_converter = MagicMock()
        with patch("importlib.util.find_spec", return_value=MagicMock()), \
             patch.object(parser, "_create_converter", return_value=mock_converter):
            result = parser._init_parser()

            assert isinstance(result, dict)
            assert result["ocr_enabled"] is True
            assert result["table_structure"] is True
            assert parser._converter is mock_converter


# =====================================================================
# Get OCR Options
# =====================================================================


@pytest.mark.unit
class TestOcrEngineSelection:
    """``OCR_ENGINE`` resolution and per-engine option building."""

    @pytest.fixture
    def settings(self):
        from application.core.settings import settings

        return settings

    def test_default_setting_is_tesseract(self):
        # Field defaults, not the live settings: a developer's ``.env`` may
        # legitimately set other engines/languages.
        from application.core.settings import Settings

        defaults = Settings.model_construct()
        assert defaults.OCR_ENGINE == "tesseract"
        assert defaults.OCR_LANGS == "eng"
        assert defaults.OCR_BACKEND == "auto"

    def test_none_reads_setting(self, settings, monkeypatch):
        from application.parser.file.docling_parser import _resolve_ocr_engine

        monkeypatch.setattr(settings, "OCR_ENGINE", "auto")
        assert _resolve_ocr_engine(None) == "auto"

    def test_unknown_engine_degrades_to_auto(self):
        from application.parser.file.docling_parser import _resolve_ocr_engine

        assert _resolve_ocr_engine("easyocr") == "auto"

    def test_tesseract_without_binary_degrades_to_auto(self, monkeypatch):
        from application.parser.file import docling_parser as dp

        monkeypatch.setattr(dp.shutil, "which", lambda name: None)
        assert dp._resolve_ocr_engine("tesseract") == "auto"

    def test_tesseract_with_binary_selected(self, monkeypatch):
        from application.parser.file import docling_parser as dp

        monkeypatch.setattr(dp.shutil, "which", lambda name: "/usr/bin/tesseract")
        assert dp._resolve_ocr_engine("tesseract") == "tesseract"

    def test_ocrmac_off_darwin_degrades_to_auto(self, monkeypatch):
        from application.parser.file import docling_parser as dp

        monkeypatch.setattr(dp.sys, "platform", "linux")
        assert dp._resolve_ocr_engine("ocrmac") == "auto"

    def test_rapidocr_missing_degrades_to_auto(self, monkeypatch):
        import sys

        from application.parser.file import docling_parser as dp

        monkeypatch.setitem(sys.modules, "rapidocr", None)
        assert dp._resolve_ocr_engine("rapidocr") == "auto"

    def test_deepseek_passes_through(self):
        from application.parser.file.docling_parser import _resolve_ocr_engine

        assert _resolve_ocr_engine("deepseek") == "deepseek"

    def test_build_auto_returns_none(self):
        from application.parser.file.docling_parser import _build_ocr_options

        assert _build_ocr_options("auto", None, True) is None

    def test_build_tesseract_reads_ocr_langs(self, settings, monkeypatch):
        pytest.importorskip("docling")
        import application.parser.file.ocr_parser as op
        from application.parser.file.docling_parser import _build_ocr_options

        # Pack inventory unknown: the resolved list is passed through untouched
        # (a host with tesseract but no chi_sim pack would otherwise drop it).
        monkeypatch.setattr(op, "tesseract_languages", lambda: None)
        monkeypatch.setattr(settings, "OCR_LANGS", "eng+chi_sim")
        options = _build_ocr_options("tesseract", None, True)

        assert type(options).__name__ == "TesseractCliOcrOptions"
        assert options.lang == ["eng", "chi_sim"]
        assert options.force_full_page_ocr is True

    def test_build_tesseract_explicit_languages_win(self, monkeypatch):
        pytest.importorskip("docling")
        import application.parser.file.ocr_parser as op
        from application.parser.file.docling_parser import _build_ocr_options

        # Pack inventory unknown: the requested list is passed through untouched.
        monkeypatch.setattr(op, "tesseract_languages", lambda: None)
        options = _build_ocr_options("tesseract", ["deu"], False)
        assert options.lang == ["deu"]

    def test_build_rapidocr(self):
        pytest.importorskip("docling")
        from application.parser.file.docling_parser import _build_ocr_options

        options = _build_ocr_options("rapidocr", None, False)
        assert type(options).__name__ == "RapidOcrOptions"
        assert options.lang == ["english"]

    def test_build_import_failure_returns_none(self, monkeypatch):
        import sys

        from application.parser.file.docling_parser import _build_ocr_options

        monkeypatch.setitem(sys.modules, "docling.datamodel.pipeline_options", None)
        assert _build_ocr_options("tesseract", ["eng"], False) is None
        assert _build_ocr_options("ocrmac", None, False) is None

    def test_build_generic_failure_returns_none(self, monkeypatch):
        import sys
        import types

        from application.parser.file.docling_parser import _build_ocr_options

        fake = types.ModuleType("docling.datamodel.pipeline_options")

        def _boom(**kwargs):
            raise RuntimeError("bad options")

        fake.RapidOcrOptions = _boom
        monkeypatch.setitem(sys.modules, "docling.datamodel.pipeline_options", fake)
        assert _build_ocr_options("rapidocr", None, False) is None


@pytest.mark.unit
class TestDeepseekVlmConverter:
    """OCR_ENGINE=deepseek swaps the whole pipeline for docling's VLM route."""

    @pytest.fixture(autouse=True)
    def _requires_docling(self):
        pytest.importorskip("docling")

    def test_deepseek_builds_vlm_converter(self, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingParser

        monkeypatch.setattr(settings, "OCR_DEEPSEEK_URL", "http://gpu-host:8000/v1/chat/completions")
        monkeypatch.setattr(settings, "OCR_DEEPSEEK_MODEL", "deepseek-ocr-x")

        built = {}

        def _capture_converter(format_options):
            built["format_options"] = format_options
            return MagicMock()

        monkeypatch.setattr(
            "docling.document_converter.DocumentConverter", _capture_converter
        )
        parser = DoclingParser(ocr_enabled=True, ocr_engine="deepseek")
        parser._create_converter()

        from docling.datamodel.base_models import InputFormat
        from docling.pipeline.vlm_pipeline import VlmPipeline

        pdf_option = built["format_options"][InputFormat.PDF]
        image_option = built["format_options"][InputFormat.IMAGE]
        assert pdf_option.pipeline_cls is VlmPipeline
        assert image_option.pipeline_cls is VlmPipeline
        vlm = pdf_option.pipeline_options.vlm_options
        assert vlm.url == "http://gpu-host:8000/v1/chat/completions"
        assert vlm.params["model"] == "deepseek-ocr-x"
        assert pdf_option.pipeline_options.enable_remote_services is True

    def test_deepseek_ignored_when_ocr_disabled(self, monkeypatch):
        from application.parser.file.docling_parser import DoclingParser

        vlm_called = []
        monkeypatch.setattr(
            DoclingParser,
            "_create_vlm_converter",
            lambda self: vlm_called.append(True),
        )
        monkeypatch.setattr("docling.document_converter.DocumentConverter", MagicMock())
        DoclingParser(ocr_enabled=False, ocr_engine="deepseek")._create_converter()

        assert vlm_called == []


# =====================================================================
# Export Content
# =====================================================================


@pytest.mark.unit
class TestExportContent:

    def test_export_markdown(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="markdown")
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = "# Title\n\nContent here"
        mock_doc.texts = []

        result = parser._export_content(mock_doc)
        assert "# Title" in result
        mock_doc.export_to_markdown.assert_called_once()

    def test_export_html(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="html")
        mock_doc = MagicMock()
        mock_doc.export_to_html.return_value = "<h1>Title</h1>"
        mock_doc.texts = []

        result = parser._export_content(mock_doc)
        assert "<h1>" in result

    def test_export_text(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="text")
        mock_doc = MagicMock()
        mock_doc.export_to_text.return_value = "Plain text content"
        mock_doc.texts = []

        result = parser._export_content(mock_doc)
        assert "Plain text" in result

    def test_fallback_to_texts_on_minimal_content(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="markdown")
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = "<!-- image -->"

        text1 = MagicMock()
        text1.text = "OCR extracted text 1"
        text2 = MagicMock()
        text2.text = "OCR extracted text 2"
        mock_doc.texts = [text1, text2]

        result = parser._export_content(mock_doc)
        assert "OCR extracted text 1" in result
        assert "OCR extracted text 2" in result

    def test_no_fallback_for_substantial_content(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="markdown")
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = "A" * 100
        mock_doc.texts = []

        result = parser._export_content(mock_doc)
        assert result == "A" * 100

    def test_fallback_skipped_when_no_texts(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="markdown")
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = "short"
        mock_doc.texts = []

        result = parser._export_content(mock_doc)
        assert result == "short"

    def test_fallback_skips_empty_texts(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser(export_format="markdown")
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = ""

        empty_text = MagicMock()
        empty_text.text = ""
        mock_doc.texts = [empty_text]

        result = parser._export_content(mock_doc)
        assert result == ""


# =====================================================================
# Parse File
# =====================================================================


@pytest.mark.unit
class TestDoclingParserParseFile:

    def test_parse_file_success(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()

        mock_converter = MagicMock()
        mock_result = MagicMock()
        mock_doc = MagicMock()
        mock_doc.export_to_markdown.return_value = "Parsed document content"
        mock_doc.texts = []
        mock_result.document = mock_doc
        mock_converter.convert.return_value = mock_result
        parser._converter = mock_converter

        result = parser.parse_file(Path("test.pdf"))
        assert "Parsed document content" in result

    def test_parse_file_inits_converter_on_first_call(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()
        parser._converter = None

        mock_converter = MagicMock()
        mock_result = MagicMock()
        mock_doc = MagicMock()
        # Long enough to clear the OCR chars-per-page floor, which a default
        # (ocr_enabled=True) parser applies to PDFs.
        mock_doc.export_to_markdown.return_value = "converted document content"
        mock_doc.texts = []
        mock_result.document = mock_doc
        mock_converter.convert.return_value = mock_result

        with patch.object(parser, "_init_parser") as mock_init:
            parser._converter = mock_converter
            mock_init.return_value = {}
            result = parser.parse_file(Path("test.pdf"))
            assert "converted document content" in result

    def test_parse_file_error_ignore_raises_instead_of_returning_error_text(self):
        """A failed conversion must never become the document's text.

        Regression: ``errors="ignore"`` used to return
        ``"[Error parsing file with docling: ...]"``, which the attachment
        worker then stored as ``attachments.content`` and handed to the LLM
        as if it were the PDF. ``errors`` controls *decoding* leniency, not
        "substitute the traceback for the document".
        """
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()
        mock_converter = MagicMock()
        mock_converter.convert.side_effect = Exception("Parse failed")
        parser._converter = mock_converter

        with pytest.raises(DocumentParseError) as excinfo:
            parser.parse_file(Path("bad.pdf"), errors="ignore")

        # The message identifies the file and preserves the cause for triage…
        assert "bad.pdf" in str(excinfo.value)
        assert "Parse failed" in str(excinfo.value)
        # …and the original exception is chained, not swallowed.
        assert isinstance(excinfo.value.__cause__, Exception)
        assert "Parse failed" in str(excinfo.value.__cause__)

    def test_parse_file_error_ignore_never_returns_a_string(self):
        """Belt-and-braces: no code path may hand back error text as content."""
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()
        mock_converter = MagicMock()
        mock_converter.convert.side_effect = RuntimeError(
            "Conversion failed for: x.pdf with status: failure. Errors: "
            "InvalidCxxCompiler: No working C++ compiler found"
        )
        parser._converter = mock_converter

        try:
            result = parser.parse_file(Path("x.pdf"), errors="ignore")
        except DocumentParseError:
            return  # expected
        pytest.fail(
            f"parse_file returned {result!r} instead of raising; error text "
            "must not be usable as document content"
        )

    def test_parse_file_error_raise(self):
        from application.parser.file.docling_parser import DoclingParser

        parser = DoclingParser()
        mock_converter = MagicMock()
        mock_converter.convert.side_effect = Exception("Parse failed")
        parser._converter = mock_converter

        with pytest.raises(Exception, match="Parse failed"):
            parser.parse_file(Path("bad.pdf"), errors="strict")


# =====================================================================
# Subclass Init
# =====================================================================


@pytest.mark.unit
class TestDoclingSubclasses:

    def test_pdf_parser_init(self):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser()
        assert parser.ocr_enabled is True
        assert parser.export_format == "markdown"

    def test_pdf_parser_custom_ocr(self):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=False, force_full_page_ocr=True)
        assert parser.ocr_enabled is False
        assert parser.force_full_page_ocr is True

    def test_docx_parser_init(self):
        from application.parser.file.docling_parser import DoclingDocxParser

        parser = DoclingDocxParser()
        assert parser.export_format == "markdown"

    def test_pptx_parser_init(self):
        from application.parser.file.docling_parser import DoclingPPTXParser

        parser = DoclingPPTXParser()
        assert parser.export_format == "markdown"

    def test_xlsx_parser_init(self):
        from application.parser.file.docling_parser import DoclingXLSXParser

        parser = DoclingXLSXParser()
        assert parser.table_structure is True

    def test_html_parser_init(self):
        from application.parser.file.docling_parser import DoclingHTMLParser

        parser = DoclingHTMLParser()
        assert parser.export_format == "markdown"

    def test_image_parser_init(self):
        from application.parser.file.docling_parser import DoclingImageParser

        parser = DoclingImageParser()
        assert parser.ocr_enabled is True
        assert parser.force_full_page_ocr is True

    def test_image_parser_custom(self):
        from application.parser.file.docling_parser import DoclingImageParser

        parser = DoclingImageParser(ocr_enabled=False)
        assert parser.ocr_enabled is False

    def test_csv_parser_init(self):
        from application.parser.file.docling_parser import DoclingCSVParser

        parser = DoclingCSVParser()
        assert parser.table_structure is True

    def test_markdown_parser_init(self):
        from application.parser.file.docling_parser import DoclingMarkdownParser

        parser = DoclingMarkdownParser()
        assert parser.export_format == "markdown"

    def test_asciidoc_parser_init(self):
        from application.parser.file.docling_parser import DoclingAsciiDocParser

        parser = DoclingAsciiDocParser()
        assert parser.export_format == "markdown"

    def test_vtt_parser_init(self):
        from application.parser.file.docling_parser import DoclingVTTParser

        parser = DoclingVTTParser()
        assert parser.export_format == "markdown"

    def test_xml_parser_init(self):
        from application.parser.file.docling_parser import DoclingXMLParser

        parser = DoclingXMLParser()
        assert parser.export_format == "markdown"


# =====================================================================
# Coverage gap tests  (lines 148-153, 289)
# =====================================================================


@pytest.mark.unit
class TestDoclingParserGaps:
    def test_csv_parser_init(self):
        """Cover line 289: DoclingCSVParser.__init__ calls super."""
        from application.parser.file.docling_parser import DoclingCSVParser

        parser = DoclingCSVParser()
        assert parser.export_format == "markdown"
        assert parser.ocr_enabled is False


@pytest.mark.unit
class TestNonOcrParsersLeaveTextAlone:
    """Office/markup docling parsers never OCR, so the tesseract CJK
    post-processing must not touch their born-digital exports."""

    @pytest.mark.parametrize(
        "name",
        [
            "DoclingDocxParser",
            "DoclingPPTXParser",
            "DoclingXLSXParser",
            "DoclingHTMLParser",
            "DoclingCSVParser",
            "DoclingMarkdownParser",
            "DoclingAsciiDocParser",
            "DoclingVTTParser",
            "DoclingXMLParser",
        ],
    )
    def test_ocr_is_off_by_construction(self, name):
        from application.parser.file import docling_parser as dp

        assert getattr(dp, name)().ocr_enabled is False

    def test_cjk_spaces_survive_in_a_docx_export(self, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingDocxParser

        monkeypatch.setattr(settings, "OCR_ENGINE", "tesseract")
        parser = DoclingDocxParser()
        # Simulate what _create_converter records for an OCR-on parser; the
        # docx parser is OCR-off so the collapse must stay inert regardless.
        parser._active_ocr_engine = "tesseract"
        text = "\u5f20 \u4e09 \u548c \u674e \u56db\n\uff21\uff22\uff23 \uff24\uff25\uff26"
        assert parser._postprocess_ocr_text(text) == text

    def test_pdf_parser_with_tesseract_still_collapses(self):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        parser._active_ocr_engine = "tesseract"
        assert parser._postprocess_ocr_text("\u4e92\u76f8 \u4fdd\u5bc6") == "\u4e92\u76f8\u4fdd\u5bc6"


# =====================================================================
# Pipeline memory caps
# =====================================================================


@pytest.mark.unit
class TestApplyPipelineCaps:
    """_apply_pipeline_caps bounds docling's threaded-pipeline buffering."""

    def test_caps_threaded_pipeline_knobs(self, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import _apply_pipeline_caps

        monkeypatch.setattr(
            settings, "DOCLING_PIPELINE_QUEUE_MAX_SIZE", 2, raising=False
        )

        class Opts:
            # docling >= 2.94 threaded pipeline — all knobs present.
            queue_max_size = 100
            layout_batch_size = 4
            table_batch_size = 4
            ocr_batch_size = 4

        opts = Opts()
        _apply_pipeline_caps(opts)

        assert opts.queue_max_size == 2
        assert opts.layout_batch_size == 1
        assert opts.table_batch_size == 1
        assert opts.ocr_batch_size == 1

    def test_queue_size_is_settings_driven(self, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import _apply_pipeline_caps

        monkeypatch.setattr(
            settings, "DOCLING_PIPELINE_QUEUE_MAX_SIZE", 6, raising=False
        )

        class Opts:
            queue_max_size = 100

        opts = Opts()
        _apply_pipeline_caps(opts)
        assert opts.queue_max_size == 6

    def test_misconfigured_zero_floors_to_one(self, monkeypatch):
        """A 0 queue depth could deadlock the threaded pipeline — floor it."""
        from application.core.settings import settings
        from application.parser.file.docling_parser import _apply_pipeline_caps

        monkeypatch.setattr(
            settings, "DOCLING_PIPELINE_QUEUE_MAX_SIZE", 0, raising=False
        )

        class Opts:
            queue_max_size = 100

        opts = Opts()
        _apply_pipeline_caps(opts)
        assert opts.queue_max_size == 1

    def test_noop_on_docling_without_threaded_pipeline(self):
        """Builds predating the threaded pipeline lack the knobs — the cap
        must be a silent no-op, not an AttributeError."""
        from application.parser.file.docling_parser import _apply_pipeline_caps

        class LegacyOpts:
            __slots__ = ("do_ocr", "do_table_structure")

            def __init__(self):
                self.do_ocr = False
                self.do_table_structure = True

        opts = LegacyOpts()
        _apply_pipeline_caps(opts)  # must not raise

        assert not hasattr(opts, "queue_max_size")
        assert not hasattr(opts, "layout_batch_size")


# =====================================================================
# Tabular size gate (CSV / XLSX)
# =====================================================================


@pytest.mark.unit
class TestDoclingTabularSizeGate:
    """Oversized tabular files must bypass docling.

    Docling materializes a ``TableCell`` object per cell (measured ~11 KB of
    RSS per 4-cell CSV row), so a multi-MB CSV balloons the worker by tens of
    GB. Above ``DOCLING_TABULAR_MAX_BYTES`` the docling tabular parsers must
    delegate to the lightweight parsers in ``tabular_parser``.
    """

    def _write_csv(self, tmp_path: Path, rows: int = 50) -> Path:
        path = tmp_path / "data.csv"
        path.write_text("\n".join(f"{i},{i * 2}" for i in range(rows)) + "\n")
        return path

    def test_oversized_csv_delegates_to_plain_csv_parser(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingCSVParser, DoclingParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 64)
        docling_parse = MagicMock(name="docling_parse")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)

        path = self._write_csv(tmp_path)
        assert path.stat().st_size > 64

        out = DoclingCSVParser().parse_file(path)

        docling_parse.assert_not_called()
        # Plain ``CSVParser`` output: rows joined with ", ", newline-separated.
        assert out.startswith("0, 0\n1, 2\n")

    def test_small_csv_still_uses_docling(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingCSVParser, DoclingParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 10_000_000)
        docling_parse = MagicMock(name="docling_parse", return_value="DOCLING")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)

        out = DoclingCSVParser().parse_file(self._write_csv(tmp_path))

        assert out == "DOCLING"
        docling_parse.assert_called_once()

    def test_gate_disabled_when_max_bytes_is_zero(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingCSVParser, DoclingParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 0)
        docling_parse = MagicMock(name="docling_parse", return_value="DOCLING")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)

        out = DoclingCSVParser().parse_file(self._write_csv(tmp_path, rows=5000))

        assert out == "DOCLING"
        docling_parse.assert_called_once()

    def test_oversized_xlsx_delegates_to_excel_parser(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingParser, DoclingXLSXParser
        from application.parser.file.tabular_parser import ExcelParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 64)
        docling_parse = MagicMock(name="docling_parse")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)
        excel_parse = MagicMock(name="excel_parse", return_value="PLAIN-XLSX")
        monkeypatch.setattr(ExcelParser, "parse_file", excel_parse)

        path = tmp_path / "big.xlsx"
        path.write_bytes(b"x" * 200)

        out = DoclingXLSXParser().parse_file(path)

        assert out == "PLAIN-XLSX"
        docling_parse.assert_not_called()
        excel_parse.assert_called_once()

    def test_small_xlsx_still_uses_docling(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import DoclingParser, DoclingXLSXParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 10_000_000)
        docling_parse = MagicMock(name="docling_parse", return_value="DOCLING")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)

        path = tmp_path / "small.xlsx"
        path.write_bytes(b"x" * 200)

        assert DoclingXLSXParser().parse_file(path) == "DOCLING"
        docling_parse.assert_called_once()


# =====================================================================
# Tabular content-size gate — XLSX compression must not defeat it
# =====================================================================


def _make_xlsx(path: Path, rows: int) -> None:
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for i in range(rows):
        ws.append([i, i * 2, i % 7, i % 13, i * 3])
    wb.save(str(path))


@pytest.mark.unit
class TestTabularContentSize:
    """XLSX is zip-compressed, so the gate must measure inner-uncompressed
    size, not on-disk bytes — otherwise a small-on-disk / many-cell xlsx
    (2.44 GB in docling) slips under a byte gate."""

    def test_xlsx_content_size_is_inner_not_ondisk(self, tmp_path):
        from application.parser.file.docling_parser import _tabular_content_size

        path = tmp_path / "data.xlsx"
        _make_xlsx(path, rows=5000)
        on_disk = path.stat().st_size
        inner = _tabular_content_size(path)
        # Repetitive numeric data compresses hard: inner XML >> zip on disk.
        assert inner > on_disk

    def test_csv_content_size_is_ondisk(self, tmp_path):
        from application.parser.file.docling_parser import _tabular_content_size

        path = tmp_path / "data.csv"
        path.write_text("a,b\n1,2\n3,4\n")
        assert _tabular_content_size(path) == path.stat().st_size

    def test_compressed_xlsx_over_inner_gate_delegates(self, tmp_path, monkeypatch):
        """The regression: on-disk < threshold < inner-uncompressed must gate."""
        from application.core.settings import settings
        from application.parser.file.docling_parser import (
            DoclingParser,
            DoclingXLSXParser,
            _tabular_content_size,
        )
        from application.parser.file.tabular_parser import ExcelParser

        path = tmp_path / "wide.xlsx"
        _make_xlsx(path, rows=5000)
        on_disk = path.stat().st_size
        inner = _tabular_content_size(path)
        assert inner > on_disk, "premise: compression hides cell count"

        # A byte gate between the two would have sent this to docling; the
        # inner-size gate must catch it.
        monkeypatch.setattr(
            settings, "DOCLING_TABULAR_MAX_BYTES", (on_disk + inner) // 2
        )
        docling_parse = MagicMock(name="docling_parse")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)
        excel_parse = MagicMock(name="excel_parse", return_value="PLAIN")
        monkeypatch.setattr(ExcelParser, "parse_file", excel_parse)

        out = DoclingXLSXParser().parse_file(path)

        assert out == "PLAIN"
        docling_parse.assert_not_called()
        excel_parse.assert_called_once()


# =====================================================================
# Markup gate (HTML / VTT) — truncate oversized element-dense markup
# =====================================================================


@pytest.mark.unit
class TestDoclingMarkupGate:
    def _write(self, tmp_path: Path, name: str, nbytes: int) -> Path:
        path = tmp_path / name
        # newline-terminated lines so the line-boundary trim has something to cut
        path.write_text(("x" * 63 + "\n") * (nbytes // 64 + 1))
        return path

    @pytest.mark.parametrize("name", ["big.html", "big.vtt"])
    def test_oversized_markup_parses_truncated_copy(self, tmp_path, monkeypatch, name):
        from application.core.settings import settings
        from application.parser.file import docling_parser as dp

        monkeypatch.setattr(settings, "DOCLING_MARKUP_MAX_BYTES", 512)
        path = self._write(tmp_path, name, 4096)
        assert path.stat().st_size > 512

        seen = {}

        def fake_parse(self_parser, file, errors="ignore"):
            p = str(file)
            seen["path"] = p
            seen["size"] = os.path.getsize(p)
            return "PARSED"

        monkeypatch.setattr(dp.DoclingParser, "parse_file", fake_parse)

        cls = dp.DoclingHTMLParser if name.endswith(".html") else dp.DoclingVTTParser
        out = cls().parse_file(path)

        assert out == "PARSED"
        assert seen["path"] != str(path), "must parse a temp copy, not the original"
        assert seen["size"] <= 512, "temp copy must be truncated to the cap"
        assert not os.path.exists(seen["path"]), "temp copy must be cleaned up"
        assert path.stat().st_size > 512, "original must be untouched"

    def test_small_markup_parses_original(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file import docling_parser as dp

        monkeypatch.setattr(settings, "DOCLING_MARKUP_MAX_BYTES", 10_000_000)
        path = self._write(tmp_path, "small.html", 1024)
        seen = {}

        def fake_parse(self_parser, file, errors="ignore"):
            seen["path"] = str(file)
            return "PARSED"

        monkeypatch.setattr(dp.DoclingParser, "parse_file", fake_parse)
        out = dp.DoclingHTMLParser().parse_file(path)

        assert out == "PARSED"
        assert seen["path"] == str(path)

    def test_markup_gate_disabled_when_zero(self, tmp_path, monkeypatch):
        from application.core.settings import settings
        from application.parser.file import docling_parser as dp

        monkeypatch.setattr(settings, "DOCLING_MARKUP_MAX_BYTES", 0)
        path = self._write(tmp_path, "big.vtt", 8192)
        seen = {}

        def fake_parse(self_parser, file, errors="ignore"):
            seen["path"] = str(file)
            return "PARSED"

        monkeypatch.setattr(dp.DoclingParser, "parse_file", fake_parse)
        dp.DoclingVTTParser().parse_file(path)

        assert seen["path"] == str(path), "disabled gate must parse the original"


# =====================================================================
# Gate seam: the oversized path through the REAL lightweight parser
# =====================================================================


@pytest.mark.unit
class TestTabularGateSeam:
    """Exercise the gate and the lightweight parser together.

    Every test in ``TestDoclingTabularSizeGate`` monkeypatches
    ``ExcelParser.parse_file`` away, so the seam between the size gate and
    the real fallback had zero coverage — which is how a crash on blank
    cells reached production and silently destroyed an upload.
    """

    def _write_xlsx_with_hole(self, path: Path) -> Path:
        openpyxl = pytest.importorskip("openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([f"col{i}" for i in range(14)])
        ws.append(list(range(14)))
        holed = list(range(14))
        holed[11] = None
        ws.append(holed)
        wb.save(path)
        return path

    def test_oversized_xlsx_with_blank_cells_parses_end_to_end(
        self, tmp_path, monkeypatch
    ):
        """The real incident, end to end: gate trips, blanks do not crash."""
        from application.core.settings import settings
        from application.parser.file.docling_parser import (
            DoclingParser,
            DoclingXLSXParser,
        )

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 64)
        docling_parse = MagicMock(name="docling_parse")
        monkeypatch.setattr(DoclingParser, "parse_file", docling_parse)

        path = self._write_xlsx_with_hole(tmp_path / "focus.xlsx")

        out = DoclingXLSXParser().parse_file(path)

        docling_parse.assert_not_called()
        assert isinstance(out, str) and out
        assert "nan" not in out

    def test_lightweight_failure_becomes_document_parse_error(
        self, tmp_path, monkeypatch
    ):
        """A fallback crash must be non-retryable and batch-skippable.

        ``DocumentParseError`` is listed in ``dont_autoretry_for`` on both
        ``ingest`` and ``store_attachment``, and is the only exception
        ``SimpleDirectoryReader.load_data`` skips rather than propagating.
        """
        from application.core.settings import settings
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import (
            DoclingParser,
            DoclingXLSXParser,
        )
        from application.parser.file.tabular_parser import ExcelParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 64)
        monkeypatch.setattr(DoclingParser, "parse_file", MagicMock())
        monkeypatch.setattr(
            ExcelParser,
            "parse_file",
            MagicMock(side_effect=TypeError("sequence item 11")),
        )

        path = self._write_xlsx_with_hole(tmp_path / "boom.xlsx")

        with pytest.raises(DocumentParseError):
            DoclingXLSXParser().parse_file(path)

    def test_oversized_csv_failure_becomes_document_parse_error(
        self, tmp_path, monkeypatch
    ):
        from application.core.settings import settings
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import (
            DoclingCSVParser,
            DoclingParser,
        )
        from application.parser.file.tabular_parser import CSVParser

        monkeypatch.setattr(settings, "DOCLING_TABULAR_MAX_BYTES", 8)
        monkeypatch.setattr(DoclingParser, "parse_file", MagicMock())
        monkeypatch.setattr(
            CSVParser, "parse_file", MagicMock(side_effect=ValueError("bad"))
        )

        path = tmp_path / "data.csv"
        path.write_text("a,b\n1,2\n" * 20)

        with pytest.raises(DocumentParseError):
            DoclingCSVParser().parse_file(path)


# =====================================================================
# torch.compile inference setting
# =====================================================================


@pytest.mark.unit
class TestApplyInferenceSettings:
    """_apply_inference_settings governs docling's torch.compile of its models.

    docling >= 2.92 compiles its layout/OCR models with torch.compile by
    default, which hard-fails TorchInductor's Metal codegen on Apple Silicon
    and buys nothing for one-shot parses.
    """

    @pytest.fixture(autouse=True)
    def _requires_docling(self):
        # These tests reach into docling.datamodel.settings for real; a base
        # install (docling is an optional extra) skips them.
        pytest.importorskip("docling")

    def test_disables_torch_compile_by_default(self, monkeypatch):
        from application.parser.file.docling_parser import _apply_inference_settings

        class Inference:
            compile_torch_models = True

        class DoclingSettings:
            inference = Inference()

        docling_settings = DoclingSettings()
        monkeypatch.setattr(
            "docling.datamodel.settings.settings", docling_settings, raising=False
        )

        _apply_inference_settings()

        assert docling_settings.inference.compile_torch_models is False

    def test_opt_in_reenables_torch_compile(self, monkeypatch):
        from application.core.settings import settings
        from application.parser.file.docling_parser import _apply_inference_settings

        monkeypatch.setattr(
            settings, "DOCLING_COMPILE_TORCH_MODELS", True, raising=False
        )

        class Inference:
            compile_torch_models = False

        class DoclingSettings:
            inference = Inference()

        docling_settings = DoclingSettings()
        monkeypatch.setattr(
            "docling.datamodel.settings.settings", docling_settings, raising=False
        )

        _apply_inference_settings()

        assert docling_settings.inference.compile_torch_models is True

    def test_noop_on_docling_without_inference_settings(self, monkeypatch):
        """Builds predating the inference settings must be a silent no-op."""
        from application.parser.file.docling_parser import _apply_inference_settings

        class DoclingSettings:
            pass

        monkeypatch.setattr(
            "docling.datamodel.settings.settings", DoclingSettings(), raising=False
        )

        _apply_inference_settings()  # must not raise

    def test_create_converter_applies_inference_settings(self, monkeypatch):
        """The cap is worthless unless the converter path actually calls it."""
        from application.parser.file.docling_parser import DoclingParser

        called = []
        monkeypatch.setattr(
            "application.parser.file.docling_parser._apply_inference_settings",
            lambda: called.append(True),
        )
        monkeypatch.setattr(
            "docling.document_converter.DocumentConverter",
            MagicMock(),
        )

        DoclingParser(ocr_enabled=False)._create_converter()

        assert called == [True]

    def test_inference_settings_applied_before_options_are_built(self, monkeypatch):
        """Ordering is load-bearing, not incidental.

        ``compile_model`` is a ``default_factory`` that reads docling's global
        at *option construction* time, so flipping the global after building
        ``PdfPipelineOptions`` is silently ignored. Pin the order.
        """
        import docling.datamodel.pipeline_options as dpo

        from application.parser.file.docling_parser import DoclingParser

        events = []
        monkeypatch.setattr(
            "application.parser.file.docling_parser._apply_inference_settings",
            lambda: events.append("settings"),
        )

        real_options = dpo.PdfPipelineOptions

        def _tracking_options(*args, **kwargs):
            events.append("options")
            return real_options(*args, **kwargs)

        monkeypatch.setattr(dpo, "PdfPipelineOptions", _tracking_options)
        monkeypatch.setattr("docling.document_converter.DocumentConverter", MagicMock())

        DoclingParser(ocr_enabled=False)._create_converter()

        assert events == ["settings", "options"], (
            "torch.compile must be disabled before PdfPipelineOptions is "
            f"constructed, got {events}"
        )


# =====================================================================
# OCR dropout guard
# =====================================================================


def _mock_conversion(markdown: str, pages: int = 1) -> MagicMock:
    """Build a mock docling ConversionResult exporting ``markdown``."""
    document = MagicMock()
    document.export_to_markdown.return_value = markdown
    document.texts = []
    document.pages = {index: object() for index in range(pages)}
    result = MagicMock()
    result.document = document
    return result


def _set_threshold(monkeypatch, value: int) -> None:
    """Point the OCR dropout guard at a specific chars-per-page floor."""
    from application.core.settings import settings as real_settings

    class _Stub:
        def __getattr__(self, name):
            return getattr(real_settings, name)

    stub = _Stub()
    stub.OCR_MIN_CHARS_PER_PAGE = value
    monkeypatch.setattr("application.core.settings.settings", stub)


@pytest.mark.unit
class TestOCRDropoutGuard:
    """A near-empty OCR parse must be loud, never stored as the document.

    In production, after a long scanned PDF, docling's pipeline degraded
    inside the worker and returned zero characters for every later scanned
    page — with no error. ``parse_file`` returned that empty string and
    ingestion reported success, so the source was indexed as empty.
    """

    def test_near_empty_first_pass_recovers_on_fresh_converter(self, caplog):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        degraded = MagicMock()
        degraded.convert.return_value = _mock_conversion(
            "<!-- image -->\n\n<!-- image -->\n\n<!-- image -->", pages=3
        )
        parser._converter = degraded

        fresh = MagicMock()
        fresh.convert.return_value = _mock_conversion(
            "Recovered scanned text for page one, page two and page three.",
            pages=3,
        )
        forced_during_retry = []

        def _create_converter():
            forced_during_retry.append(parser.force_full_page_ocr)
            return fresh

        with patch.object(
            parser, "_create_converter", side_effect=_create_converter
        ) as create, caplog.at_level("WARNING"):
            result = parser.parse_file(Path("scan.pdf"))

        assert "Recovered scanned text" in result
        # Exactly one retry, on a converter that did not exist before.
        assert create.call_count == 1
        assert fresh.convert.call_count == 1
        assert degraded.convert.call_count == 1
        # The retry forces full-page OCR...
        assert forced_during_retry == [True]
        # ...but that must not stick to the parser or its converter, or every
        # later file in the worker pays for full-page OCR it does not need.
        assert parser.force_full_page_ocr is False
        assert parser._converter is None
        messages = " ".join(caplog.messages)
        assert "scan.pdf" in messages
        assert "near-empty" in messages
        assert "Recovered scan.pdf on retry" in messages

    def test_near_empty_both_passes_raises_document_parse_error(self):
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        degraded = MagicMock()
        degraded.convert.return_value = _mock_conversion("<!-- image -->", pages=100)
        parser._converter = degraded

        fresh = MagicMock()
        fresh.convert.return_value = _mock_conversion("", pages=100)

        with patch.object(parser, "_create_converter", return_value=fresh):
            with pytest.raises(DocumentParseError) as excinfo:
                parser.parse_file(Path("scan.pdf"))

        message = str(excinfo.value)
        assert "scan.pdf" in message
        assert "OCR produced 0 chars over 100 pages" in message
        assert "OCR pipeline dropout" in message
        assert "not indexed" in message
        # The dropout diagnosis must survive: the generic handler must not
        # re-wrap it into "Failed to parse ... with docling".
        assert "Failed to parse" not in message
        assert parser._converter is None

    def test_dropout_error_reports_an_unread_text_layer(self, monkeypatch):
        """A text layer docling ignored points at the pipeline, not the scan."""
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingPDFParser

        monkeypatch.setattr(
            "application.parser.file.docling_parser._pdf_text_layer_probe",
            lambda file: (12, 48_000),
        )
        parser = DoclingPDFParser(ocr_enabled=True)
        parser._converter = MagicMock()
        parser._converter.convert.return_value = _mock_conversion("", pages=12)

        with patch.object(parser, "_create_converter", return_value=parser._converter):
            with pytest.raises(DocumentParseError) as excinfo:
                parser.parse_file(Path("scan.pdf"))

        assert "The PDF carries a 48000-char text layer" in str(excinfo.value)

    def test_ocr_disabled_returns_near_empty_content_untouched(self):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=False)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion("tiny", pages=40)
        parser._converter = converter

        with patch.object(parser, "_create_converter") as create:
            assert parser.parse_file(Path("text.pdf")) == "tiny"

        assert create.call_count == 0
        assert converter.convert.call_count == 1
        assert parser._converter is converter

    def test_threshold_zero_disables_the_guard(self, monkeypatch):
        from application.parser.file.docling_parser import DoclingPDFParser

        _set_threshold(monkeypatch, 0)
        parser = DoclingPDFParser(ocr_enabled=True)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion("", pages=100)
        parser._converter = converter

        with patch.object(parser, "_create_converter") as create:
            assert parser.parse_file(Path("scan.pdf")) == ""

        assert create.call_count == 0
        assert parser._converter is converter

    def test_healthy_ocr_parse_converts_once(self):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion(
            "A page of genuinely scanned text. " * 10, pages=2
        )
        parser._converter = converter

        with patch.object(parser, "_create_converter") as create:
            result = parser.parse_file(Path("scan.pdf"))

        assert "genuinely scanned text" in result
        assert converter.convert.call_count == 1
        assert create.call_count == 0
        assert parser._converter is converter

    def test_threshold_is_per_page_not_per_document(self):
        """200 chars is healthy for one page and suspicious for a hundred."""
        from application.parser.file.docling_parser import DoclingPDFParser

        content = "x" * 200

        parser = DoclingPDFParser(ocr_enabled=True)
        parser._converter = MagicMock()
        parser._converter.convert.return_value = _mock_conversion(content, pages=1)
        with patch.object(parser, "_create_converter") as create:
            assert parser.parse_file(Path("one-page.pdf")) == content
        assert create.call_count == 0

        # Over a hundred pages the same text is below the floor, so the guard
        # spends a full-page-OCR retry on it. It still indexes afterwards --
        # sparse is not the same as dropped (see the dropout tests below).
        long_doc = DoclingPDFParser(ocr_enabled=True)
        long_doc._converter = MagicMock()
        long_doc._converter.convert.return_value = _mock_conversion(content, pages=100)
        with patch.object(
            long_doc, "_create_converter", return_value=long_doc._converter
        ) as create, patch(
            "application.parser.file.docling_parser._pdf_text_layer_probe",
            return_value=(100, 0),
        ):
            assert long_doc.parse_file(Path("hundred-page.pdf")) == content
        assert create.call_count == 1

    def test_guard_does_not_apply_to_non_ocr_formats(self):
        """DOCX parsers construct with OCR off: they never OCR anything."""
        from application.parser.file.docling_parser import DoclingDocxParser

        parser = DoclingDocxParser()
        assert parser.ocr_enabled is False
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion("Hi", pages=1)
        parser._converter = converter

        with patch.object(parser, "_create_converter") as create:
            assert parser.parse_file(Path("memo.docx")) == "Hi"

        assert create.call_count == 0

    def test_guard_applies_to_images(self):
        """The retry covers images -- a degraded converter may still recover one."""
        from application.parser.file.docling_parser import DoclingImageParser

        parser = DoclingImageParser(ocr_enabled=True)
        degraded = MagicMock()
        degraded.convert.return_value = _mock_conversion("<!-- image -->")
        parser._converter = degraded

        fresh = MagicMock()
        fresh.convert.return_value = _mock_conversion("Recovered caption text.")

        with patch.object(parser, "_create_converter", return_value=fresh) as create:
            assert parser.parse_file(Path("scan.png")) == "Recovered caption text."

        assert create.call_count == 1

    def test_a_text_free_image_indexes_rather_than_failing_the_upload(self, caplog):
        """A chart, logo or photo has no text to find; that is not a dropout.

        DocumentParseError is in ``dont_autoretry_for``, so raising here fails a
        single-file upload permanently.
        """
        from application.parser.file.docling_parser import DoclingImageParser

        parser = DoclingImageParser(ocr_enabled=True)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion("<!-- image -->")
        parser._converter = converter

        with patch.object(parser, "_create_converter", return_value=converter):
            with caplog.at_level(logging.WARNING):
                assert parser.parse_file(Path("chart.png")) == "<!-- image -->"

        assert "text-sparse" in caplog.text

    def test_a_text_sparse_scan_indexes_rather_than_failing_the_upload(self):
        """20 pages of pictures with a few captions is the document, not a fault."""
        from application.parser.file.docling_parser import DoclingPDFParser

        content = "Fig 1. Fig 2. Fig 3."
        parser = DoclingPDFParser(ocr_enabled=True)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion(content, pages=20)
        parser._converter = converter

        with patch.object(
            parser, "_create_converter", return_value=converter
        ), patch(
            "application.parser.file.docling_parser._pdf_text_layer_probe",
            return_value=(20, 0),
        ):
            assert parser.parse_file(Path("catalog.pdf")) == content

    def test_a_multi_page_zero_char_parse_is_still_a_dropout(self):
        """The incident this guard exists for: every page OCR'd to nothing."""
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion("", pages=40)
        parser._converter = converter

        with patch.object(
            parser, "_create_converter", return_value=converter
        ), patch(
            "application.parser.file.docling_parser._pdf_text_layer_probe",
            return_value=(40, 0),
        ):
            with pytest.raises(DocumentParseError, match="scan.pdf"):
                parser.parse_file(Path("scan.pdf"))

    def test_a_pdf_whose_text_layer_was_missed_is_still_a_dropout(self):
        """Text docling should have read without OCR at all -- positive evidence."""
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        converter = MagicMock()
        converter.convert.return_value = _mock_conversion("stub", pages=20)
        parser._converter = converter

        with patch.object(
            parser, "_create_converter", return_value=converter
        ), patch(
            "application.parser.file.docling_parser._pdf_text_layer_probe",
            return_value=(20, 5000),
        ):
            with pytest.raises(DocumentParseError, match="text layer"):
                parser.parse_file(Path("report.pdf"))

    def test_retry_conversion_failure_is_still_a_parse_error(self):
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        parser._converter = MagicMock()
        parser._converter.convert.return_value = _mock_conversion("", pages=2)

        fresh = MagicMock()
        fresh.convert.side_effect = RuntimeError("OCR model unavailable")

        with patch.object(parser, "_create_converter", return_value=fresh):
            with pytest.raises(DocumentParseError, match="OCR model unavailable"):
                parser.parse_file(Path("scan.pdf"))

        assert parser.force_full_page_ocr is False
        assert parser._converter is None


@pytest.mark.unit
class TestOCRDropoutHelpers:

    def test_image_placeholders_are_not_text(self):
        from application.parser.file.docling_parser import _text_char_count

        assert _text_char_count(None) == 0
        assert _text_char_count("") == 0
        assert _text_char_count("<!-- image -->\n\n<!-- image -->") == 0
        assert _text_char_count("  <!-- image -->  abc  ") == 3

    def test_page_count_prefers_document_pages(self):
        from application.parser.file.docling_parser import _result_page_count

        result = _mock_conversion("x", pages=7)
        assert _result_page_count(result, Path("a.pdf")) == 7

    def test_page_count_falls_back_to_num_pages(self):
        from application.parser.file.docling_parser import _result_page_count

        result = _mock_conversion("x", pages=0)
        result.document.num_pages.return_value = 5
        assert _result_page_count(result, Path("a.pdf")) == 5

    def test_page_count_falls_back_to_result_pages(self):
        from application.parser.file.docling_parser import _result_page_count

        result = _mock_conversion("x", pages=0)
        result.document.num_pages.return_value = None
        result.pages = [object(), object()]
        assert _result_page_count(result, Path("a.pdf")) == 2

    def test_page_count_defaults_to_one(self):
        from application.parser.file.docling_parser import _result_page_count

        result = _mock_conversion("x", pages=0)
        result.document.num_pages.return_value = None
        result.pages = []
        assert _result_page_count(result, Path("a.pdf")) == 1

    def test_images_are_always_one_page(self):
        from application.parser.file.docling_parser import _result_page_count

        result = _mock_conversion("x", pages=9)
        assert _result_page_count(result, Path("a.png")) == 1

    def test_threshold_reads_settings_with_a_default(self, monkeypatch):
        from application.parser.file.docling_parser import _ocr_min_chars_per_page

        assert _ocr_min_chars_per_page() == 20
        _set_threshold(monkeypatch, 5)
        assert _ocr_min_chars_per_page() == 5
        _set_threshold(monkeypatch, "nonsense")
        assert _ocr_min_chars_per_page() == 20

    def test_text_layer_probe_survives_an_unreadable_pdf(self, tmp_path):
        from application.parser.file.docling_parser import _pdf_text_layer_probe

        assert _pdf_text_layer_probe(tmp_path / "missing.pdf") == (0, 0)
        broken = tmp_path / "broken.pdf"
        broken.write_bytes(b"not a pdf")
        assert _pdf_text_layer_probe(broken) == (0, 0)


@pytest.mark.unit
class TestForceFullPageOCRWiring:
    """``force_full_page_ocr`` must reach the pipeline for any OCR engine."""

    @pytest.fixture(autouse=True)
    def _requires_docling(self):
        # Builds real ``docling.datamodel.pipeline_options``; a base install
        # (docling is an optional extra) skips these.
        pytest.importorskip("docling")

    def _pipeline_options(self, monkeypatch, **kwargs):
        import docling.datamodel.pipeline_options as dpo

        from application.parser.file.docling_parser import DoclingParser

        built = []
        real_options = dpo.PdfPipelineOptions

        def _tracking_options(*args, **opts):
            options = real_options(*args, **opts)
            built.append(options)
            return options

        monkeypatch.setattr(dpo, "PdfPipelineOptions", _tracking_options)
        monkeypatch.setattr("docling.document_converter.DocumentConverter", MagicMock())
        DoclingParser(**kwargs)._create_converter()
        return built[0]

    def test_forced_with_default_auto_options(self, monkeypatch):
        options = self._pipeline_options(
            monkeypatch,
            ocr_enabled=True,
            ocr_engine="auto",
            force_full_page_ocr=True,
        )
        assert options.ocr_options.force_full_page_ocr is True

    def test_not_forced_by_default(self, monkeypatch):
        options = self._pipeline_options(
            monkeypatch,
            ocr_enabled=True,
            ocr_engine="auto",
            force_full_page_ocr=False,
        )
        assert options.ocr_options.force_full_page_ocr is False


@pytest.mark.unit
class TestDoclingTesseractPostprocessing:
    """docling's tesseract path gets the same CJK glyph-space cleanup as the native engine."""

    def _parser_with_export(self, engine, ocr_enabled=True):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=ocr_enabled)
        parser._active_ocr_engine = engine
        document = MagicMock()
        document.export_to_markdown.return_value = "## 互相 保密 协议\n\nhello world 你 好"
        return parser, document

    def test_tesseract_output_collapses_cjk_spaces(self):
        parser, document = self._parser_with_export("tesseract")
        assert parser._export_content(document) == "## 互相保密协议\n\nhello world 你好"

    def test_other_engines_untouched(self):
        parser, document = self._parser_with_export("deepseek")
        assert parser._export_content(document) == "## 互相 保密 协议\n\nhello world 你 好"

    def test_ocr_off_untouched(self):
        parser, document = self._parser_with_export("tesseract", ocr_enabled=False)
        assert parser._export_content(document) == "## 互相 保密 协议\n\nhello world 你 好"


@pytest.mark.unit
class TestDoclingOcrPages:
    def test_converts_each_requested_page_with_page_range(self, tmp_path):
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        converter = MagicMock()

        def convert(source, page_range=None):
            result = MagicMock()
            result.document.export_to_markdown.return_value = f"page {page_range[0]} text " * 5
            return result

        converter.convert.side_effect = convert
        parser._converter = converter
        pdf = tmp_path / "doc.pdf"
        pdf.write_bytes(b"%PDF-1.4")

        texts = parser.ocr_pages(pdf, [2, 0])

        assert set(texts) == {2, 0}
        assert texts[2].startswith("page 3 text")
        assert texts[0].startswith("page 1 text")
        assert [c.kwargs["page_range"] for c in converter.convert.call_args_list] == [(3, 3), (1, 1)]

    def test_failure_is_a_parse_error(self, tmp_path):
        from application.parser.file.base_parser import DocumentParseError
        from application.parser.file.docling_parser import DoclingPDFParser

        parser = DoclingPDFParser(ocr_enabled=True)
        parser._converter = MagicMock()
        parser._converter.convert.side_effect = RuntimeError("boom")
        pdf = tmp_path / "doc.pdf"
        pdf.write_bytes(b"%PDF-1.4")
        with pytest.raises(DocumentParseError, match="page 1 of doc.pdf"):
            parser.ocr_pages(pdf, [0])


@pytest.mark.unit
class TestTesseractLanguageFilter:
    def test_uninstalled_packs_are_dropped_with_a_warning(self, monkeypatch, caplog):
        pytest.importorskip("docling")
        import application.parser.file.ocr_parser as op
        from application.parser.file.docling_parser import _build_ocr_options

        monkeypatch.setattr(op, "tesseract_languages", lambda: frozenset({"eng", "osd"}))
        with caplog.at_level("WARNING"):
            options = _build_ocr_options("tesseract", ["eng", "chi_sim"], False)
        assert options.lang == ["eng"]
        assert "chi_sim" in caplog.text

    def test_all_packs_missing_falls_back_to_eng(self, monkeypatch):
        pytest.importorskip("docling")
        import application.parser.file.ocr_parser as op
        from application.parser.file.docling_parser import _build_ocr_options

        monkeypatch.setattr(op, "tesseract_languages", lambda: frozenset({"eng"}))
        assert _build_ocr_options("tesseract", ["xyz"], False).lang == ["eng"]

    def test_unknown_inventory_keeps_the_list(self, monkeypatch):
        pytest.importorskip("docling")
        import application.parser.file.ocr_parser as op
        from application.parser.file.docling_parser import _build_ocr_options

        monkeypatch.setattr(op, "tesseract_languages", lambda: None)
        assert _build_ocr_options("tesseract", ["eng", "deu"], False).lang == ["eng", "deu"]
